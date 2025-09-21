import openpyxl
from openpyxl.styles import Font, Alignment
import os

def write_test_results_excel(results, filename="test_results.xlsx", sheet_name="Test Results"):
    # Nếu file chưa tồn tại → tạo mới
    if not os.path.exists(filename):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name
        ws.append(["Time", "Test Name", "Email", "Password", "Expected", "Actual", "Status", "Screenshot"])

        # Style header
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        wb.save(filename)

    # Ghi dữ liệu
    wb = openpyxl.load_workbook(filename)
    # Nếu sheet tồn tại → dùng lại, nếu không thì tạo mới
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(sheet_name)
        ws.append(["Time", "Test Name", "Email", "Password", "Expected", "Actual", "Status", "Screenshot"])

    for result in results:
        ws.append([
            result["Time"],
            result["Test Name"],
            result["Email"],
            result["Password"],
            result["Expected"],
            result["Actual"],
            result["Status"],
            result["Screenshot"]
        ])

    wb.save(filename)
